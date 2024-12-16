import pandas as pd
import openpyxl
import os

class file:
    #将目标文件视为活动文件
    def __init__(self,path):
        self.path = path
        try:
            self.wb = openpyxl.load_workbook(self.path)
            self.ws = self.wb.active
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = 'Sheet1'
            # 如果是新文件，写入第一行标题
            headers = ['文献名称', '参考文献']
            self.ws.append(headers)

    #在文件末行添加数据
    def write_endrow(self,data):
        # 找到最后一行的下一个空行
        next_row = self.ws.max_row + 1

        # 填充数据
        for row in data:
            self.ws.append(row)

        # 保存文件
        self.wb.save(self.path)
        
    # def judge_file_exist(is_check,check_times,check_interval_sec,check_path,check_ext):
    #     """
    #     检测函数
    #     :param is_check:是否检测True表示检测False 不检查
    #     :param check_times:检测次数
    #     :param check_interval:检测时间间隔（默认值）
    #     :param check_path:检测路径（默认值）
    #     :param chedk_ext:检测扩展名（默认值）
    #     :return:返回真假
    #     """
    #     if os.path.exists(check_path) is False:
    #         raise Exception("路径不存在...")
    #     if str.isdigit(str(check_times)) is False:
    #         raise Exception("检测次数不是数字...")
    #     if is_check is False:
    #         return True,"不进行检测" #直接执行下一步操作
    #     else:
    #         for number in range(1,int(check_times)):
    #             print("正在进行第" str(number) "次检测...")
    #             files = os.listdir(check_path) # 读取目录下所有文件

    #             file_number=len(files)
    #             if file_number ==0:
    #                 #不存在任何文件，休眠一会 继续执行下一次
    #                 sleep(int(check_interval_sec))  # 休眠一会
    #                 continue
    #             elif file_number==1:
    #                 file_name=files[0]
    #                 #crdownload
    #                 file_full_name=check_path os.sep file_name
    #                 file_ext=os.path.splitext(file_full_name)[-1]
    #                 if "crdownload"==str(file_ext.split(".")[1]):
    #                     sleep(int(check_interval_sec))  # 休眠一会
    #                     continue
    #                 if "tmp" == str(file_ext.split(".")[1]):
    #                     sleep(int(check_interval_sec))  # 休眠一会
    #                     continue
    #                 for e in check_ext.split("|"):
    #                     if e == str(file_ext.split(".")[1]):
    #                         return True,file_full_name
    #                     else:
    #                         sleep(int(check_interval_sec))  # 休眠一会
    #                         continue
    #             else:
    #                 #多个文件认定是False直接执行下一步操作
    #                 for file in files:
    #                     file_ext = str(os.path.splitext(file)[-1])
    #                     if file_ext==".rar":
    #                         os.unlink(file)
    #                     if file_ext==".zip":
    #                         os.unlink(file)
    #                 return False,"多个文件认定是False"
    #         return False,"可能是不存在文件或者是tmp或crdownload文件"
        